"""Local concrete-test PDF processor.

Run this file on the computer that holds the PDF folder. It processes PDFs
one at a time and creates Concrete_Test_Log.xlsx. The PDFs never enter the
Streamlit browser app, so the source folder may be larger than 1 GB.
"""

import argparse
import gc
import re
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FILENAME_PATTERN = re.compile(
    r"^W02229_(?:Concrete Test|Test)_"
    r"Sample (?P<sample_id>\d{3,4})_"
    r"(?P<report_date>\d{4}-\d{2}-\d{2})_"
    r"Report (?P<report_id>\d{6}-\d{2})_"
    r"Compressive Strength of Concrete_"
    r"(?:\d{6}|\d{8})\.pdf$",
    flags=re.IGNORECASE,
)

COLUMNS = [
    "Report Date",
    "Sample ID",
    "Location Details",
    "Air Temp",
    "Concrete Temp",
    "Slump",
    "Air Content",
    "Min Temp",
    "Max Temp",
    "7 Day Test Average",
    "28 Day Test Average",
]


def clean_value(value):
    if value is None:
        return ""
    value = str(value).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", value).strip()


def clean_number(value):
    if value is None:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", clean_value(value).replace(",", ""))
    if not match:
        return None
    number = float(match.group(0))
    return int(number) if number.is_integer() else number


def extract_pattern(text, patterns):
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            return clean_value(match.group(1))
    return ""


def parse_mixed_number(value):
    text = clean_value(value)
    match = re.search(r"(\d+)\s*[- ]\s*(\d+)\s*/\s*(\d+)", text)
    if match:
        whole, numerator, denominator = map(int, match.groups())
        if denominator:
            return round(whole + numerator / denominator, 3)
    return clean_number(text)


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", " ", clean_value(value).lower()).strip()


def find_lab_header_row(table):
    for index, row in enumerate(table):
        combined = " | ".join(normalize_header(cell) for cell in (row or []))
        if "test age days" in combined and "strength psi" in combined:
            return index
    return None


def extract_pdf_once(pdf_path):
    text_parts = []
    strengths_by_age = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                text_parts.append(page_text)

            if "strength" not in page_text.lower():
                continue

            for table in page.extract_tables():
                if not table:
                    continue
                header_index = find_lab_header_row(table)
                if header_index is None:
                    continue

                header_map = {}
                for column_index, heading in enumerate(table[header_index]):
                    heading = normalize_header(heading)
                    if "test age days" in heading:
                        header_map["age"] = column_index
                    elif "strength psi" in heading:
                        header_map["strength"] = column_index

                if "age" not in header_map or "strength" not in header_map:
                    continue

                for row in table[header_index + 1:]:
                    if not row:
                        continue
                    age_index = header_map["age"]
                    strength_index = header_map["strength"]
                    age = clean_number(row[age_index] if age_index < len(row) else "")
                    strength = clean_number(
                        row[strength_index] if strength_index < len(row) else ""
                    )
                    if age is not None and strength is not None:
                        strengths_by_age.setdefault(age, []).append(strength)

    return "\n".join(text_parts), strengths_by_age


def extract_location(text):
    value = extract_pattern(
        text,
        [
            r"Location\s+Details\s*:\s*(.+?)(?=\n|$)",
            r"Location\s+Details\s*\n\s*(.+?)(?=\n|$)",
            r"Pour\s+Description\s*:\s*(.+?)(?=\n|$)",
        ],
    )
    for label in ["Mix Design", "Supplier", "Technician", "Weather", "Field Measurements"]:
        value = re.split(
            rf"\s+{re.escape(label)}\s*:", value, maxsplit=1, flags=re.IGNORECASE
        )[0]
    return clean_value(value)


def extract_measurements(text):
    air_temp = extract_pattern(text, [
        r"Air\s+Temperature\s*(?:\(F\))?\s*:\s*(-?\d+(?:\.\d+)?)",
        r"Air\s+Temp(?:erature)?\s*(?:\(F\))?\s*:\s*(-?\d+(?:\.\d+)?)",
    ])
    concrete_temp = extract_pattern(text, [
        r"Concrete\s+Temp(?:erature)?\s*(?:\(F\))?\s*:\s*(-?\d+(?:\.\d+)?)",
    ])
    slump = extract_pattern(text, [
        r"Slump\s*(?:\(in\))?\s*:\s*(\d+\s*[- ]\s*\d+\s*/\s*\d+)",
        r"Slump\s*(?:\(in\))?\s*:\s*(\d+(?:\.\d+)?)",
    ])
    air_content = extract_pattern(text, [
        r"Air\s+Content\s*(?:\(%\))?\s*:\s*(-?\d+(?:\.\d+)?)",
    ])
    min_max = extract_pattern(text, [
        r"Min\s*/\s*Max\s+Temp\s*(?:\(F\))?\s*:\s*"
        r"(-?\d+(?:\.\d+)?\s*/\s*-?\d+(?:\.\d+)?)",
    ])

    min_temp = max_temp = None
    if min_max:
        pieces = re.split(r"\s*/\s*", min_max, maxsplit=1)
        if len(pieces) == 2:
            min_temp = clean_number(pieces[0])
            max_temp = clean_number(pieces[1])

    return [
        clean_number(air_temp),
        clean_number(concrete_temp),
        parse_mixed_number(slump),
        clean_number(air_content),
        min_temp,
        max_temp,
    ]


def reported_average(text, age):
    return clean_number(extract_pattern(text, [
        rf"{age}\s*Day\s*-\s*([\d,]+(?:\.\d+)?)",
        rf"{age}\s*Day\s*Average\s*:?\s*([\d,]+(?:\.\d+)?)",
    ]))


def average_for_age(strengths, text, age):
    values = strengths.get(age, [])
    if values:
        return round(sum(values) / len(values))
    return reported_average(text, age)


def process_pdf(pdf_path):
    match = FILENAME_PATTERN.fullmatch(pdf_path.name)
    if not match:
        return None, "filename did not match"

    info = match.groupdict()
    datetime.strptime(info["report_date"], "%Y-%m-%d")
    text, strengths = extract_pdf_once(pdf_path)

    row = [
        info["report_date"],
        info["sample_id"],
        extract_location(text),
        *extract_measurements(text),
        average_for_age(strengths, text, 7),
        average_for_age(strengths, text, 28),
    ]

    del text
    del strengths
    gc.collect()
    return row, None


def create_workbook(path):
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Concrete Test Log")
    sheet.append(COLUMNS)
    workbook.save(path)


def append_row(path, row):
    workbook = load_workbook(path)
    sheet = workbook["Concrete Test Log"]
    sheet.append(row)
    workbook.save(path)
    workbook.close()


def format_workbook(path):
    workbook = load_workbook(path)
    sheet = workbook["Concrete Test Log"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [15, 14, 52, 12, 16, 12, 14, 12, 12, 22, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row_number in range(2, sheet.max_row + 1):
        sheet[f"A{row_number}"].number_format = "yyyy-mm-dd"
        for column in ["D", "E", "F", "G", "H", "I"]:
            sheet[f"{column}{row_number}"].number_format = "0.00"
        sheet[f"J{row_number}"].number_format = "#,##0"
        sheet[f"K{row_number}"].number_format = "#,##0"

    workbook.save(path)
    workbook.close()


def main():
    parser = argparse.ArgumentParser(description="Build the Concrete Test Log locally.")
    parser.add_argument("folder", help="Folder containing concrete-test PDFs")
    parser.add_argument(
        "--output",
        default="Concrete_Test_Log.xlsx",
        help="Output Excel path",
    )
    args = parser.parse_args()

    folder = Path(args.folder)
    output = Path(args.output)
    pdf_files = sorted(path for path in folder.rglob("*.pdf") if path.is_file())

    if not pdf_files:
        raise SystemExit(f"No PDFs found in {folder}")

    create_workbook(output)
    successful = 0
    skipped = []

    for index, pdf_path in enumerate(pdf_files, start=1):
        try:
            row, error = process_pdf(pdf_path)
            if row is None:
                skipped.append((pdf_path.name, error))
            else:
                append_row(output, row)
                successful += 1
        except Exception as exc:
            skipped.append((pdf_path.name, f"{type(exc).__name__}: {exc}"))

        print(f"Processed {index}/{len(pdf_files)}: {pdf_path.name}")

    format_workbook(output)
    print(f"Created: {output.resolve()}")
    print(f"Successful reports: {successful}")
    print(f"Skipped reports: {len(skipped)}")

    if skipped:
        skipped_path = output.with_name("Skipped_PDFs.txt")
        skipped_path.write_text(
            "\n".join(f"{name}: {reason}" for name, reason in skipped),
            encoding="utf-8",
        )
        print(f"Skipped-file details: {skipped_path.resolve()}")


if __name__ == "__main__":
    main()
