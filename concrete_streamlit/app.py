import gc
import io
import re
import zipfile
from uuid import uuid4
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import pdfplumber
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 1. STREAMLIT AND RESOURCE SETTINGS
# ============================================================

st.set_page_config(
    page_title="Concrete Test Dashboard",
    page_icon="📊",
    layout="wide",
)

# These limits protect Streamlit Community Cloud from large memory spikes.
MAX_PDF_FILES = 100
MAX_TOTAL_UPLOAD_MB = 150
MAX_PDF_MB = 25

# Completed workbooks are stored on temporary disk, not in session memory.
RESULT_DIRECTORY = Path("/tmp/concrete_test_dashboard_results")
RESULT_DIRECTORY.mkdir(parents=True, exist_ok=True)

# Accept either "Concrete Test" or "Test" in the filename.
FILENAME_PATTERN = re.compile(
    r"^W02229_(?:Concrete Test|Test)_"
    r"Sample (?P<sample_id>\d{3,4})_"
    r"(?P<report_date>\d{4}-\d{2}-\d{2})_"
    r"Report (?P<report_id>\d{6}-\d{2})_"
    r"Compressive Strength of Concrete_"
    r"(?:\d{6}|\d{8})\.pdf$",
    flags=re.IGNORECASE,
)

MASTER_COLUMNS = [
    "Report Date",
    "Sample ID",
    "Location Details",
    "Air Temp",
    "Concrete Temp",
    "Slump",
    "Air Content",
    "Min Temp",
    "Max Temp",
    "7 Day Test Average",
    "28 Day Test Average",
]

CHART_FIELDS = [
    "Air Temp",
    "Concrete Temp",
    "Slump",
    "Air Content",
    "Min Temp",
    "Max Temp",
]


# ============================================================
# 2. GENERAL TEXT AND NUMBER HELPERS
# ============================================================


def clean_value(value):
    """Remove line breaks and repeated spaces."""
    if value is None:
        return ""

    value = str(value).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", value).strip()


def clean_number(value):
    """Return the first numeric value as an int or float."""
    if value is None:
        return None

    text = clean_value(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)

    if not match:
        return None

    number = float(match.group(0))
    return int(number) if number.is_integer() else number


def extract_pattern(text, patterns):
    """Return the first captured value from several patterns."""
    for pattern in patterns:
        match = re.search(
            pattern,
            text,
            flags=re.IGNORECASE | re.MULTILINE,
        )

        if match:
            return clean_value(match.group(1))

    return ""


def parse_mixed_number(value):
    """Convert a mixed slump value such as 7-1/2 into 7.5."""
    text = clean_value(value)

    mixed_match = re.search(
        r"(\d+)\s*[- ]\s*(\d+)\s*/\s*(\d+)",
        text,
    )

    if mixed_match:
        whole, numerator, denominator = map(int, mixed_match.groups())

        if denominator:
            return round(whole + numerator / denominator, 3)

    fraction_match = re.fullmatch(
        r"\s*(\d+)\s*/\s*(\d+)\s*",
        text,
    )

    if fraction_match:
        numerator, denominator = map(int, fraction_match.groups())

        if denominator:
            return round(numerator / denominator, 3)

    return clean_number(text)


def normalize_header(value):
    """Normalize a PDF table heading for matching."""
    value = clean_value(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


# ============================================================
# 3. PDF FIELD EXTRACTION
# ============================================================


def extract_location_details(text):
    """Extract the pour description shown beside Location Details."""
    location = extract_pattern(
        text,
        [
            r"Location\s+Details\s*:\s*(.+?)(?=\n|$)",
            r"Location\s+Details\s*\n\s*(.+?)(?=\n|$)",
            r"Location\s+Detail\s*:\s*(.+?)(?=\n|$)",
            r"Pour\s+Description\s*:\s*(.+?)(?=\n|$)",
        ],
    )

    for label in [
        "Mix Design",
        "Supplier",
        "Technician",
        "Weather",
        "Field Measurements",
        "On-Site Admixtures",
    ]:
        location = re.split(
            rf"\s+{re.escape(label)}\s*:",
            location,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]

    return clean_value(location)


def extract_field_measurements(text):
    """Extract requested field measurements without units."""
    air_temp = extract_pattern(
        text,
        [
            r"Air\s+Temperature\s*(?:\(F\))?\s*:\s*(-?\d+(?:\.\d+)?)",
            r"Air\s+Temp(?:erature)?\s*(?:\(F\))?\s*:\s*(-?\d+(?:\.\d+)?)",
        ],
    )

    concrete_temp = extract_pattern(
        text,
        [
            r"Concrete\s+Temp(?:erature)?\s*(?:\(F\))?\s*:\s*(-?\d+(?:\.\d+)?)",
        ],
    )

    slump = extract_pattern(
        text,
        [
            r"Slump\s*(?:\(in\))?\s*:\s*(\d+\s*[- ]\s*\d+\s*/\s*\d+)",
            r"Slump\s*(?:\(in\))?\s*:\s*(\d+\s*/\s*\d+)",
            r"Slump\s*(?:\(in\))?\s*:\s*(\d+(?:\.\d+)?)",
        ],
    )

    air_content = extract_pattern(
        text,
        [
            r"Air\s+Content\s*(?:\(%\))?\s*:\s*(-?\d+(?:\.\d+)?)",
        ],
    )

    min_max = extract_pattern(
        text,
        [
            r"Min\s*/\s*Max\s+Temp\s*(?:\(F\))?\s*:\s*"
            r"(-?\d+(?:\.\d+)?\s*/\s*-?\d+(?:\.\d+)?)",
        ],
    )

    min_temp = None
    max_temp = None

    if min_max:
        pieces = re.split(r"\s*/\s*", min_max, maxsplit=1)

        if len(pieces) == 2:
            min_temp = clean_number(pieces[0])
            max_temp = clean_number(pieces[1])

    return {
        "Air Temp": clean_number(air_temp),
        "Concrete Temp": clean_number(concrete_temp),
        "Slump": parse_mixed_number(slump),
        "Air Content": clean_number(air_content),
        "Min Temp": min_temp,
        "Max Temp": max_temp,
    }


def find_lab_header_row(table):
    """Find the lab row containing Test Age Days and Strength PSI."""
    for row_index, row in enumerate(table):
        combined = " | ".join(
            normalize_header(cell)
            for cell in (row or [])
        )

        if "test age days" in combined and "strength psi" in combined:
            return row_index

    return None


def extract_pdf_content_once(pdf_source):
    """
    Open a PDF once and extract both text and break strengths.

    The full PDF text exists only while this function runs. After the output
    row is written to Excel, the text and PDF buffer are deleted.
    """
    pdf_source.seek(0)
    text_parts = []
    strengths_by_age = {}

    with pdfplumber.open(pdf_source) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()

            if page_text:
                text_parts.append(page_text)

            for table in page.extract_tables():
                if not table:
                    continue

                header_row_index = find_lab_header_row(table)

                if header_row_index is None:
                    continue

                header_map = {}

                for column_index, heading in enumerate(
                    table[header_row_index]
                ):
                    normalized = normalize_header(heading)

                    if "test age days" in normalized:
                        header_map["age"] = column_index
                    elif "strength psi" in normalized:
                        header_map["strength"] = column_index

                if "age" not in header_map or "strength" not in header_map:
                    continue

                for row in table[header_row_index + 1:]:
                    if not row:
                        continue

                    age_index = header_map["age"]
                    strength_index = header_map["strength"]

                    age_value = row[age_index] if age_index < len(row) else ""
                    strength_value = (
                        row[strength_index]
                        if strength_index < len(row)
                        else ""
                    )

                    age = clean_number(age_value)
                    strength = clean_number(strength_value)

                    if age is not None and strength is not None:
                        strengths_by_age.setdefault(age, []).append(strength)

    pdf_source.seek(0)
    return "\n".join(text_parts), strengths_by_age


def extract_reported_average(text, test_age):
    """Read a reported test-age average if table rows are unavailable."""
    value = extract_pattern(
        text,
        [
            rf"{test_age}\s*Day\s*-\s*([\d,]+(?:\.\d+)?)",
            rf"{test_age}\s*Day\s*Average\s*:?\s*([\d,]+(?:\.\d+)?)",
            rf"{test_age}[-\s]*Day\s*Average\s*:?\s*([\d,]+(?:\.\d+)?)",
        ],
    )

    return clean_number(value)


def calculate_test_average(strengths_by_age, text, test_age):
    """Average all completed breaks at a requested test age."""
    strengths = strengths_by_age.get(test_age, [])

    if strengths:
        return round(sum(strengths) / len(strengths))

    return extract_reported_average(text, test_age)


def process_pdf_to_row(pdf_source, filename):
    """Convert one PDF directly into one Excel row."""
    match = FILENAME_PATTERN.fullmatch(filename)

    if not match:
        return None, "Filename does not match an accepted convention."

    filename_info = match.groupdict()

    try:
        datetime.strptime(filename_info["report_date"], "%Y-%m-%d")
    except ValueError:
        return None, "Filename contains an invalid report date."

    text, strengths_by_age = extract_pdf_content_once(pdf_source)
    measurements = extract_field_measurements(text)

    row = [
        filename_info["report_date"],
        filename_info["sample_id"],
        extract_location_details(text),
        measurements["Air Temp"],
        measurements["Concrete Temp"],
        measurements["Slump"],
        measurements["Air Content"],
        measurements["Min Temp"],
        measurements["Max Temp"],
        calculate_test_average(strengths_by_age, text, 7),
        calculate_test_average(strengths_by_age, text, 28),
    ]

    # Explicitly release extracted content before returning.
    del text
    del strengths_by_age
    del measurements

    return row, None


# ============================================================
# 4. STREAMING EXCEL WORKBOOK
# ============================================================


def create_workbook(output_path):
    """Create the master workbook before PDF processing starts."""
    workbook = Workbook(write_only=False)
    worksheet = workbook.active
    worksheet.title = "Concrete Test Log"
    worksheet.append(MASTER_COLUMNS)
    workbook.save(output_path)
    workbook.close()


def append_row_to_excel(output_path, row):
    """
    Append one extracted row to Excel, save it, and close the workbook.

    This prevents the app from retaining a growing list of extracted PDF
    records while the batch is being processed.
    """
    workbook = load_workbook(output_path)
    worksheet = workbook["Concrete Test Log"]
    worksheet.append(row)
    workbook.save(output_path)
    workbook.close()


def format_finished_workbook(output_path):
    """Apply final Excel styling after every PDF row has been written."""
    workbook = load_workbook(output_path)
    worksheet = workbook["Concrete Test Log"]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 38

    widths = [15, 14, 52, 12, 16, 12, 14, 12, 12, 22, 24]

    for column_number, width in enumerate(widths, start=1):
        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = width

    for row_number in range(2, worksheet.max_row + 1):
        worksheet[f"A{row_number}"].number_format = "yyyy-mm-dd"

        for column_letter in ["D", "E", "F", "G", "H", "I"]:
            worksheet[
                f"{column_letter}{row_number}"
            ].number_format = "0.00"

        worksheet[f"J{row_number}"].number_format = "#,##0"
        worksheet[f"K{row_number}"].number_format = "#,##0"

    workbook.save(output_path)
    workbook.close()


def load_compact_dataframe(output_path):
    """
    Load only the small final Excel table needed for display and charts.

    Raw PDFs and extracted full text are no longer retained at this point.
    """
    dataframe = pd.read_excel(
        output_path,
        sheet_name="Concrete Test Log",
        engine="openpyxl",
    )

    dataframe["Report Date"] = pd.to_datetime(
        dataframe["Report Date"],
        errors="coerce",
    )

    dataframe.sort_values(
        ["Report Date", "Sample ID"],
        inplace=True,
        na_position="last",
    )
    dataframe.reset_index(drop=True, inplace=True)

    return dataframe


# ============================================================
# 5. FILE COUNTING AND SEQUENTIAL PROCESSING
# ============================================================


def get_upload_size_bytes(uploaded_file):
    """Return an upload's byte size without making a second copy."""
    size = getattr(uploaded_file, "size", None)

    if size is not None:
        return size

    current_position = uploaded_file.tell()
    uploaded_file.seek(0, io.SEEK_END)
    size = uploaded_file.tell()
    uploaded_file.seek(current_position)
    return size


def count_pdf_items(uploaded_items):
    """Count PDF items from uploads without extracting ZIP members."""
    total_count = 0
    notes = []

    for item in uploaded_items:
        lower_name = item.name.lower()

        if lower_name.endswith(".pdf"):
            total_count += 1
            continue

        if not lower_name.endswith(".zip"):
            notes.append((item.name, "Unsupported file type."))
            continue

        try:
            item.seek(0)

            with zipfile.ZipFile(item) as archive:
                count = sum(
                    1
                    for member in archive.infolist()
                    if (
                        not member.is_dir()
                        and member.filename.lower().endswith(".pdf")
                    )
                )

                total_count += count
                notes.append(
                    (item.name, f"Found {count} PDF file(s) in this ZIP.")
                )

        except zipfile.BadZipFile:
            notes.append((item.name, "Could not read this ZIP file."))
        finally:
            item.seek(0)

    return total_count, notes


def process_all_uploads(uploaded_items, output_path, total_pdf_count, progress):
    """
    Process direct PDFs and ZIP members one at a time.

    A ZIP member is read, converted into one Excel row, and immediately
    released before the next member is opened.
    """
    processed_count = 0
    successful_count = 0
    processing_messages = []

    for item in uploaded_items:
        lower_name = item.name.lower()

        if lower_name.endswith(".pdf"):
            pdf_size_mb = get_upload_size_bytes(item) / 1024 / 1024

            if pdf_size_mb > MAX_PDF_MB:
                processed_count += 1
                processing_messages.append(
                    (
                        item.name,
                        f"Skipped because the PDF is {pdf_size_mb:.1f} MB.",
                    )
                )
            else:
                try:
                    row, error = process_pdf_to_row(item, item.name)

                    if row is not None:
                        append_row_to_excel(output_path, row)
                        successful_count += 1
                        del row
                    else:
                        processing_messages.append((item.name, error))

                except Exception as exc:
                    processing_messages.append(
                        (item.name, f"{type(exc).__name__}: {exc}")
                    )

                processed_count += 1

            progress.progress(
                processed_count / max(total_pdf_count, 1),
                text=f"Processed {processed_count} of {total_pdf_count} PDFs",
            )

            gc.collect()
            continue

        if not lower_name.endswith(".zip"):
            continue

        try:
            item.seek(0)

            with zipfile.ZipFile(item) as archive:
                pdf_members = [
                    member
                    for member in archive.infolist()
                    if (
                        not member.is_dir()
                        and member.filename.lower().endswith(".pdf")
                    )
                ]

                for member in pdf_members:
                    filename = Path(member.filename).name
                    member_size_mb = member.file_size / 1024 / 1024

                    if member_size_mb > MAX_PDF_MB:
                        processing_messages.append(
                            (
                                filename,
                                (
                                    f"Skipped because the PDF is "
                                    f"{member_size_mb:.1f} MB."
                                ),
                            )
                        )
                    else:
                        pdf_buffer = None

                        try:
                            # Only this one ZIP member is held in memory.
                            pdf_buffer = io.BytesIO(archive.read(member))
                            row, error = process_pdf_to_row(
                                pdf_buffer,
                                filename,
                            )

                            if row is not None:
                                append_row_to_excel(output_path, row)
                                successful_count += 1
                                del row
                            else:
                                processing_messages.append((filename, error))

                        except Exception as exc:
                            processing_messages.append(
                                (filename, f"{type(exc).__name__}: {exc}")
                            )

                        finally:
                            if pdf_buffer is not None:
                                pdf_buffer.close()
                            del pdf_buffer

                    processed_count += 1
                    progress.progress(
                        processed_count / max(total_pdf_count, 1),
                        text=(
                            f"Processed {processed_count} of "
                            f"{total_pdf_count} PDFs"
                        ),
                    )
                    gc.collect()

        except zipfile.BadZipFile:
            processing_messages.append(
                (item.name, "Could not read this ZIP file.")
            )
        finally:
            item.seek(0)

    return successful_count, processing_messages


# ============================================================
# 6. ONE-CHART-AT-A-TIME QUADRATIC ANALYSIS
# ============================================================


def add_quadratic_series(
    figure,
    dataframe,
    x_field,
    y_field,
    series_name,
    color,
    symbol,
):
    """Add scatter points and a second-order polynomial curve."""
    plot_data = dataframe[
        [x_field, y_field, "Sample ID", "Location Details"]
    ].dropna()

    if plot_data.empty:
        return

    figure.add_trace(
        go.Scatter(
            x=plot_data[x_field],
            y=plot_data[y_field],
            mode="markers",
            name=series_name,
            marker={
                "color": color,
                "size": 9,
                "symbol": symbol,
            },
            customdata=plot_data[["Sample ID", "Location Details"]],
            hovertemplate=(
                "Sample ID: %{customdata[0]}<br>"
                "Location: %{customdata[1]}<br>"
                f"{x_field}: %{{x}}<br>"
                f"{y_field}: %{{y:,.0f}} PSI<extra></extra>"
            ),
        )
    )

    if len(plot_data) < 3 or plot_data[x_field].nunique() < 3:
        return

    x_values = plot_data[x_field].astype(float).to_numpy()
    y_values = plot_data[y_field].astype(float).to_numpy()

    coefficient_a, coefficient_b, intercept = np.polyfit(
        x_values,
        y_values,
        2,
    )

    x_line = np.linspace(x_values.min(), x_values.max(), 150)
    y_line = (
        coefficient_a * x_line ** 2
        + coefficient_b * x_line
        + intercept
    )

    y_predicted = (
        coefficient_a * x_values ** 2
        + coefficient_b * x_values
        + intercept
    )

    residual_sum = np.sum((y_values - y_predicted) ** 2)
    total_sum = np.sum((y_values - np.mean(y_values)) ** 2)
    r_squared = 1 - residual_sum / total_sum if total_sum else np.nan

    equation = (
        f"y = {coefficient_a:.3f}x² "
        f"+ {coefficient_b:.2f}x "
        f"+ {intercept:.1f}"
    )
    r_text = (
        f"R² = {r_squared:.3f}"
        if not np.isnan(r_squared)
        else "R² unavailable"
    )

    figure.add_trace(
        go.Scatter(
            x=x_line,
            y=y_line,
            mode="lines",
            name=f"{series_name} quadratic fit ({equation}; {r_text})",
            line={"color": color, "width": 3},
            hoverinfo="skip",
        )
    )


def build_selected_chart(dataframe, x_field):
    """Create only the chart currently selected by the user."""
    figure = go.Figure()

    add_quadratic_series(
        figure,
        dataframe,
        x_field,
        "7 Day Test Average",
        "7 Day Average",
        "#4472C4",
        "circle",
    )

    add_quadratic_series(
        figure,
        dataframe,
        x_field,
        "28 Day Test Average",
        "28 Day Average",
        "#ED7D31",
        "diamond",
    )

    figure.update_layout(
        title=f"{x_field} vs Concrete Compressive Strength",
        template="plotly_white",
        height=560,
        hovermode="closest",
        legend_title="Test Age and Quadratic Curve",
        xaxis={
            "title": f"{x_field} (numeric field measurement)",
            "showgrid": True,
            "gridcolor": "#D9D9D9",
            "tickformat": ".2f",
            "nticks": 12,
            "zeroline": False,
        },
        yaxis={
            "title": "Average Concrete Compressive Strength (PSI)",
            "showgrid": True,
            "gridcolor": "#D9D9D9",
            "tickformat": ",.0f",
            "nticks": 12,
            "rangemode": "tozero",
        },
    )

    return figure


# ============================================================
# 7. RESULT FILE AND SESSION HELPERS
# ============================================================


def remove_result_file(result_path):
    """Delete a prior temporary workbook from server disk."""
    if not result_path:
        return

    try:
        path = Path(result_path)
        if path.exists():
            path.unlink()
    except OSError:
        pass


def reset_upload_widget():
    """Change the uploader key so Streamlit releases uploaded file objects."""
    st.session_state["uploader_generation"] += 1


def clear_current_result():
    """Delete workbook, clear session references, and reset the uploader."""
    remove_result_file(st.session_state.get("result_path"))

    generation = st.session_state.get("uploader_generation", 0) + 1

    for session_key in list(st.session_state.keys()):
        del st.session_state[session_key]

    st.session_state["uploader_generation"] = generation
    st.cache_data.clear()
    st.cache_resource.clear()
    gc.collect()


def read_workbook_for_download(result_path):
    """Read workbook bytes only when the download button is selected."""
    return Path(result_path).read_bytes()


# ============================================================
# 8. STREAMLIT USER INTERFACE
# ============================================================

if "uploader_generation" not in st.session_state:
    st.session_state["uploader_generation"] = 0

if "result_path" not in st.session_state:
    st.session_state["result_path"] = ""

if "processing_details" not in st.session_state:
    st.session_state["processing_details"] = []

st.title("Concrete Test Dashboard")
st.caption(
    "Each PDF is processed once, its row is written directly to Excel, and "
    "the uploaded PDF is released automatically before results are displayed."
)

with st.expander("How to use this tool", expanded=True):
    st.markdown(
        """
        **Step 1: Prepare the reports**
        - Keep the original PDF names.
        - Accepted names can contain either `Concrete Test` or `Test` after `W02229_`.
        - Sample IDs may contain 3 or 4 digits, including Sample 922.
        - Compress a complete folder into a `.zip` file before uploading it.

        **Step 2: Upload and process**
        - Upload individual PDFs, ZIP folders, or a mixture.
        - Select **Process reports and build Excel**.
        - The app reads one PDF at a time and writes one row directly to Excel.

        **Step 3: Automatic PDF removal**
        - After processing finishes, the app resets the upload widget automatically.
        - Uploaded PDF and ZIP objects are no longer kept in the active page session.
        - Only the compact Excel workbook remains on temporary server disk.

        **Step 4: Review, download, and chart**
        - Review the extracted master table.
        - Download the formatted `.xlsx` workbook.
        - Select one field measurement at a time for quadratic strength analysis.

        **Step 5: Clear everything**
        - Use **Delete workbook and clear memory** at the bottom when finished.
        - This removes the temporary Excel file, table, chart data, uploader state, and caches.
        """
    )

st.subheader("Step 1: Upload PDF reports or ZIP folders")

uploaded_items = st.file_uploader(
    "Drop ZIP folders or concrete test PDFs here",
    type=["pdf", "zip"],
    accept_multiple_files=True,
    key=f"pdf_upload_{st.session_state['uploader_generation']}",
    help=(
        "Accepted examples include W02229_Test_Sample 922_... and "
        "W02229_Concrete Test_Sample 1826_...."
    ),
)

process_button = st.button(
    "Process reports and build Excel",
    type="primary",
    disabled=not uploaded_items,
    use_container_width=True,
)

if process_button:
    # Delete the prior result before building a replacement workbook.
    remove_result_file(st.session_state.get("result_path"))
    st.session_state["result_path"] = ""
    st.session_state["processing_details"] = []
    gc.collect()

    total_upload_bytes = sum(
        get_upload_size_bytes(item)
        for item in uploaded_items
    )
    total_upload_mb = total_upload_bytes / 1024 / 1024

    if total_upload_mb > MAX_TOTAL_UPLOAD_MB:
        st.error(
            f"This upload is {total_upload_mb:.1f} MB. "
            f"Please keep each batch below {MAX_TOTAL_UPLOAD_MB} MB."
        )
        st.stop()

    total_pdf_count, count_notes = count_pdf_items(uploaded_items)

    if total_pdf_count > MAX_PDF_FILES:
        st.error(
            f"This upload contains {total_pdf_count} PDFs. "
            f"Please upload no more than {MAX_PDF_FILES} PDFs at one time."
        )
        st.stop()

    if total_pdf_count == 0:
        st.error("No PDF files were found in the upload.")
        st.stop()

    st.subheader("Step 2: Stream each report directly into Excel")
    progress = st.progress(
        0,
        text="Preparing the Excel workbook...",
    )

    output_path = RESULT_DIRECTORY / f"Concrete_Test_Log_{uuid4().hex}.xlsx"
    create_workbook(output_path)

    successful_count, processing_messages = process_all_uploads(
        uploaded_items,
        output_path,
        total_pdf_count,
        progress,
    )

    progress.empty()

    # Close each original Streamlit UploadedFile after extraction.
    for uploaded_item in uploaded_items:
        try:
            uploaded_item.close()
        except Exception:
            pass

    if successful_count == 0:
        remove_result_file(output_path)
        st.session_state["processing_details"] = count_notes + processing_messages
        reset_upload_widget()
        gc.collect()
        st.rerun()

    format_finished_workbook(output_path)

    # Store only a small path string and processing messages in session state.
    st.session_state["result_path"] = str(output_path)
    st.session_state["processing_details"] = count_notes + processing_messages

    # This key change is essential. It drops the old uploader widget and its PDFs.
    reset_upload_widget()
    del uploaded_items
    gc.collect()
    st.rerun()

result_path = st.session_state.get("result_path", "")
result_exists = bool(result_path) and Path(result_path).exists()

if not result_exists:
    st.info(
        "Upload reports and select Process reports and build Excel. "
        "The uploader will clear itself automatically after processing."
    )

    if st.session_state.get("processing_details"):
        with st.expander("Processing details"):
            for name, message in st.session_state["processing_details"]:
                st.write(f"{name}: {message}")

    st.stop()

st.success(
    "Processing complete. Uploaded PDF and ZIP objects were removed from the "
    "active uploader session. The finished workbook is stored temporarily."
)

# Load only the compact final table. Raw PDFs and full text are already gone.
master_df = load_compact_dataframe(result_path)

st.subheader("Step 3: Review the Concrete Test Log")

metric_one, metric_two, metric_three = st.columns(3)
metric_one.metric("Reports", len(master_df))
metric_two.metric(
    "7-Day Results",
    int(master_df["7 Day Test Average"].notna().sum()),
)
metric_three.metric(
    "28-Day Results",
    int(master_df["28 Day Test Average"].notna().sum()),
)

st.dataframe(
    master_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Report Date": st.column_config.DateColumn(
            "Report Date",
            format="YYYY-MM-DD",
        ),
        "7 Day Test Average": st.column_config.NumberColumn(
            "7 Day Test Average",
            format="%d",
        ),
        "28 Day Test Average": st.column_config.NumberColumn(
            "28 Day Test Average",
            format="%d",
        ),
    },
)

st.subheader("Step 4: Download the Excel workbook")

st.download_button(
    "Download Concrete Test Log (.xlsx)",
    data=lambda: read_workbook_for_download(result_path),
    file_name="Concrete_Test_Log.xlsx",
    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),
    type="primary",
)

if st.session_state.get("processing_details"):
    with st.expander("Processing details"):
        for name, message in st.session_state["processing_details"]:
            st.write(f"{name}: {message}")

st.divider()
st.header("Step 5: Explore one strength chart at a time")
st.caption(
    "Blue circles and blue quadratic curves represent 7-day averages. "
    "Orange diamonds and orange quadratic curves represent 28-day averages."
)

selected_chart = st.selectbox(
    "Choose the field measurement for the horizontal axis",
    options=CHART_FIELDS,
)

selected_figure = build_selected_chart(master_df, selected_chart)

st.plotly_chart(
    selected_figure,
    use_container_width=True,
)

del selected_figure
gc.collect()

st.divider()
st.subheader("Memory control")
st.warning(
    "After downloading the workbook, use the button below to delete the "
    "temporary workbook and clear the remaining table and chart memory."
)

if st.button(
    "Delete workbook and clear memory",
    type="secondary",
    use_container_width=True,
):
    clear_current_result()
    st.rerun()
